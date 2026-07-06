###############################################################################
# 1. Module Level Documentation
###############################################################################
"""
Stock Tracker — business layer (Excel inventory + distributor APIs).

No GUI code in this module. The PySide6 layer in `src/gui/` uses `StockTracker`.
See `docs/especificacao/PROJETO_STOCKTRACKER.md` for data model and SCAN flow.
"""

###############################################################################
# 2. Imports
###############################################################################
import re
import sys
import threading
from datetime import date, datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from .suppliers import search_part, supplier_label
from .suppliers.base import SupplierId

###############################################################################
# 3. Constants and Global Variables
###############################################################################
_SUPPLIER_SEARCH_ORDER: tuple[SupplierId, ...] = (
    "mouser",
    "tme",
    "rs",
    "digikey",
    "robert_mauser",
)

PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from config.credentials import load_secrets

from .app_settings import get_low_stock_threshold

DATA_DIR = PROJECT_ROOT / "data"
EXCEL_FILE = DATA_DIR / "stock.xlsx"

SHEET_COMPONENTS = "Components"
SHEET_GENERIC = "Generic"
SHEET_MASSIVE_LEGACY = "Massive"  # legacy Excel sheet name (auto-renamed)
SHEET_HISTORY = "History"
SHEET_EQUIPMENTS = "Equipments"
SHEET_MATERIALS_LEGACY = "Materials"  # legacy Excel sheet name (migration only)

# Equipments sheet columns (1-based Excel column index)
EQ_COL_ID = 1
EQ_COL_SUPPLIER_REF = 2
EQ_COL_SERIAL = 3
EQ_COL_NAME = 4
EQ_COL_DESCRIPTION = 5
EQ_COL_CALIB_DATE = 6
EQ_COL_CALIB_EXPIRY = 7
EQ_COL_DATASHEET = 8
EQ_COL_IMAGE = 9
EQ_COL_LOCATION = 10
EQ_COL_LOANED = 11
EQ_COL_LOANED_TO = 12
EQ_COL_LOAN_PLACE = 13
EQ_COL_LOAN_SINCE = 14

EQUIPMENT_IMAGE_SEPARATOR = ";"

SHEET_EQUIPMENT_LOANS = "EquipmentLoans"

COMP_COL_STOCK = 7
COMP_COL_LOCATION = 8

COMPONENT_LOCATION_SEPARATOR = ";"

_LOW_STOCK_FONT = Font(color="FF0000", bold=True)
_NORMAL_STOCK_FONT = Font(color="000000", bold=False)

# Generic sheet columns (resistors, capacitors — high-volume passives)
MASSIVE_COL_ID = 1
MASSIVE_COL_TYPE = 2
MASSIVE_COL_VALUE = 3
MASSIVE_COL_TOLERANCE = 4
MASSIVE_COL_PACKAGE = 5
MASSIVE_COL_NAME = 6
MASSIVE_COL_STOCK = 7
MASSIVE_COL_SUPPLIER_REF = 8
MASSIVE_COL_DIELECTRIC = 9
MASSIVE_COL_VOLTAGE = 10
MASSIVE_COL_NOTES = 11
MASSIVE_COL_LOCATION = 12

_MASSIVE_TYPES = frozenset({"R", "C"})

_CATALOG_LOOKUP_LOCK = threading.Lock()
_MIN_PARTIAL_REF_LEN = 4  # avoid "X" matching "RMH05-DK-XX" via substring

###############################################################################
# 4. StockTracker class
###############################################################################


class StockTracker:
    """
    Class:
        Manages component inventory in Excel and optional distributor catalog lookup.
    Args:
        excel_path (Path | None): Override path to workbook; default `data/stock.xlsx`.
        api_key (str): Optional Mouser key; otherwise loaded from `config/secrets.py`.
    Example:
        tracker = StockTracker()
        wb = tracker.get_workbook()
        sheet = tracker.get_components_sheet(wb)
        row = tracker.find_component(sheet, "581-SR4M3DC12")
    """

    def __init__(self, excel_path: Optional[Path] = None, api_key: str = ""):
        self.excel_file = Path(excel_path) if excel_path else EXCEL_FILE
        self._secrets = load_secrets()
        # Compatibilidade: api_key no construtor ou MOUSER_API_KEY em secrets
        if api_key:
            self._secrets["MOUSER_API_KEY"] = api_key
        self.api_key = str(self._secrets.get("MOUSER_API_KEY", "")).strip()
        self._catalog_session_cache: dict[str, dict] = {}
        self._ensure_data_folder()

    def _ensure_data_folder(self) -> None:
        DATA_DIR.mkdir(parents=True, exist_ok=True)

    # ------------------------------------------------------------------
    # Excel helpers
    # ------------------------------------------------------------------
    def get_workbook(self):
        try:
            return load_workbook(self.excel_file)
        except FileNotFoundError:
            wb = Workbook()
            wb.active.title = SHEET_COMPONENTS
            return wb

    def get_components_sheet(self, workbook):
        if SHEET_COMPONENTS not in workbook.sheetnames:
            sheet = workbook.create_sheet(SHEET_COMPONENTS)
        else:
            sheet = workbook[SHEET_COMPONENTS]

        if sheet.max_row == 1 and sheet["A1"].value is None:
            sheet.append([
                "ID", "Supplier Reference", "Manufacturer",
                "Manufacturer Reference", "Value", "Description", "Stock",
                "Location",
            ])
        self._ensure_column_header(sheet, COMP_COL_LOCATION, "Location")
        return sheet

    def get_massive_sheet(self, workbook):
        """Generic inventory sheet (resistors / capacitors)."""
        if (
            SHEET_GENERIC not in workbook.sheetnames
            and SHEET_MASSIVE_LEGACY in workbook.sheetnames
        ):
            workbook[SHEET_MASSIVE_LEGACY].title = SHEET_GENERIC
        if SHEET_GENERIC not in workbook.sheetnames:
            sheet = workbook.create_sheet(SHEET_GENERIC)
        else:
            sheet = workbook[SHEET_GENERIC]

        if sheet.max_row == 1 and sheet["A1"].value is None:
            sheet.append([
                "ID",
                "Type",
                "Value",
                "Tolerance",
                "Package",
                "Name",
                "Stock",
                "Supplier Reference",
                "Dielectric",
                "Voltage",
                "Notes",
                "Location",
            ])
        self._ensure_column_header(sheet, MASSIVE_COL_LOCATION, "Location")
        return sheet

    def get_history_sheet(self, workbook):
        if SHEET_HISTORY not in workbook.sheetnames:
            history = workbook.create_sheet(SHEET_HISTORY)
        else:
            history = workbook[SHEET_HISTORY]

        if history.max_row == 1 and history["A1"].value is None:
            history.append([
                "Date", "User", "Supplier Reference",
                "Movement", "Quantity", "Stock After",
            ])
        return history

    @staticmethod
    def _stock_cell_int(value) -> int | None:
        if value is None or str(value).strip() == "":
            return 0
        try:
            return int(value)
        except (TypeError, ValueError):
            return None

    def _apply_low_stock_formatting(self, workbook) -> None:
        """Highlight Stock cells in red when at or below the configured threshold."""
        threshold = self.low_stock_threshold
        if SHEET_COMPONENTS in workbook.sheetnames:
            sheet = workbook[SHEET_COMPONENTS]
            for row_idx in range(2, sheet.max_row + 1):
                if self.row_is_empty(
                    tuple(sheet.iter_rows(min_row=row_idx, max_row=row_idx))[0]
                ):
                    continue
                cell = sheet.cell(row=row_idx, column=COMP_COL_STOCK)
                stock = self._stock_cell_int(cell.value)
                cell.font = (
                    _LOW_STOCK_FONT
                    if stock is not None and stock <= threshold
                    else _NORMAL_STOCK_FONT
                )

        if SHEET_GENERIC in workbook.sheetnames:
            sheet = workbook[SHEET_GENERIC]
            for row_idx in range(2, sheet.max_row + 1):
                row = tuple(sheet.iter_rows(min_row=row_idx, max_row=row_idx))[0]
                if self.row_is_empty(row) or self._massive_row_is_header(row):
                    continue
                cell = sheet.cell(row=row_idx, column=MASSIVE_COL_STOCK)
                stock = self._stock_cell_int(cell.value)
                cell.font = (
                    _LOW_STOCK_FONT
                    if stock is not None and stock <= threshold
                    else _NORMAL_STOCK_FONT
                )

    def save_workbook(self, workbook) -> bool:
        try:
            from .excel_backups import backup_excel_file

            self._apply_low_stock_formatting(workbook)
            backup_excel_file(self.excel_file)
            workbook.save(self.excel_file)
            return True
        except PermissionError:
            print("ERRO: Fecha o stock.xlsx no Excel antes de guardar.")
            return False

    def ensure_workbook_sheets(self) -> bool:
        """Create Components, Equipments and History sheets if missing, then save."""
        workbook = self.get_workbook()
        self.get_components_sheet(workbook)
        self.get_massive_sheet(workbook)
        self.get_history_sheet(workbook)
        self.get_equipments_sheet(workbook)
        self.get_equipment_loans_sheet(workbook)
        self._order_workbook_sheets(workbook)
        self._remove_default_sheet(workbook)
        return self.save_workbook(workbook)

    @staticmethod
    def _order_workbook_sheets(workbook) -> None:
        order = (
            SHEET_COMPONENTS,
            SHEET_GENERIC,
            SHEET_EQUIPMENTS,
            SHEET_EQUIPMENT_LOANS,
            SHEET_HISTORY,
        )
        for target_idx, name in enumerate(order):
            if name not in workbook.sheetnames:
                continue
            current_idx = workbook.sheetnames.index(name)
            offset = target_idx - current_idx
            if offset:
                workbook.move_sheet(name, offset)

    @staticmethod
    def _remove_default_sheet(workbook) -> None:
        if "Sheet" not in workbook.sheetnames:
            return
        sheet = workbook["Sheet"]
        if sheet.max_row <= 1 and not sheet["A1"].value:
            workbook.remove(sheet)

    @staticmethod
    def _ensure_column_header(sheet, column: int, header: str) -> None:
        current = str(sheet.cell(row=1, column=column).value or "").strip()
        if not current:
            sheet.cell(row=1, column=column, value=header)

    # ------------------------------------------------------------------
    # Search / barcode
    # ------------------------------------------------------------------
    @staticmethod
    def extract_part_number(text: str) -> str:
        text = text.strip()
        match = re.search(r"P([A-Z0-9\-]+)Q", text, re.IGNORECASE)
        if match:
            return match.group(1)
        return text

    @staticmethod
    def normalize_ref(text) -> str:
        if text is None:
            return ""
        value = str(text).strip().upper()
        return value

    def refs_match(
        self, query: str, *candidates: str, allow_short_substring: bool = False
    ) -> bool:
        """Match part refs or text; partial substring needs min length unless description."""
        q = self.normalize_ref(query)
        if not q or len(q) < 2:
            return False
        for raw in candidates:
            c = self.normalize_ref(raw)
            if not c:
                continue
            if q == c:
                return True
            if allow_short_substring:
                if q in c or c in q:
                    return True
            elif len(q) >= _MIN_PARTIAL_REF_LEN and len(c) >= _MIN_PARTIAL_REF_LEN:
                if q in c or c in q:
                    return True
        return False

    def row_matches_search(self, query: str, row) -> bool:
        """Excel row match: strict on refs, looser on description."""
        if self.refs_match(
            query,
            row[1].value,
            row[2].value,
            row[3].value,
        ):
            return True
        return self.refs_match(query, row[5].value, allow_short_substring=True)

    @staticmethod
    def row_is_empty(row) -> bool:
        return all(
            cell.value is None or str(cell.value).strip() == ""
            for cell in row
        )

    def find_component(self, sheet, part_number: str):
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            if self.row_is_empty(row):
                continue
            if self.refs_match(
                part_number,
                row[1].value,
                row[2].value,
                row[3].value,
            ):
                return row
            if self.refs_match(
                part_number, row[5].value, allow_short_substring=True
            ):
                return row
        return None

    def find_component_any(self, sheet, *queries: str):
        seen = set()
        for query in queries:
            key = self.normalize_ref(query)
            if not key or key in seen:
                continue
            seen.add(key)
            row = self.find_component(sheet, query)
            if row:
                return row
        return None

    def find_component_by_supplier_ref(self, sheet, supplier_reference: str):
        """Exact match on supplier reference column (column B) — for manual add/edit."""
        ref_n = self.normalize_ref(supplier_reference)
        if not ref_n:
            return None
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            if self.row_is_empty(row):
                continue
            if self.normalize_ref(row[1].value) == ref_n:
                return row
        return None

    def search_in_excel_all(self, sheet, query: str) -> list:
        """All rows matching query (Mouser ref, manufacturer, mfr ref, description)."""
        results = []
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            if self.row_is_empty(row):
                continue
            if self.row_matches_search(query, row):
                results.append(row)
        return results

    def search_in_excel(self, sheet, query: str):
        matches = self.search_in_excel_all(sheet, query)
        return matches[0] if matches else None

    def excel_autocomplete_terms(self, sheet) -> list[str]:
        """Referencias no Excel para sugestoes ao escrever (pesquisa manual)."""
        seen: set[str] = set()
        terms: list[str] = []
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            if self.row_is_empty(row):
                continue
            for col in (1, 2, 3):
                text = str(row[col].value or "").strip()
                if len(text) < 2:
                    continue
                key = text.upper()
                if key in seen:
                    continue
                seen.add(key)
                terms.append(text)
        return sorted(terms, key=lambda x: x.upper())

    def excel_autocomplete_supplier_refs(self, sheet) -> list[str]:
        """Supplier references (column 1) for barcode/scan autocomplete."""
        seen: set[str] = set()
        refs: list[str] = []
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            if self.row_is_empty(row):
                continue
            text = str(row[1].value or "").strip()
            if len(text) < 1:
                continue
            key = text.upper()
            if key in seen:
                continue
            seen.add(key)
            refs.append(text)
        return sorted(refs, key=lambda x: x.upper())

    def excel_autocomplete_mouser_refs(self, sheet) -> list[str]:
        """Backward-compatible alias."""
        return self.excel_autocomplete_supplier_refs(sheet)

    @staticmethod
    def part_supplier_reference(part: dict, fallback: str = "") -> str:
        return str(
            part.get("supplier_part_number")
            or part.get("MouserPartNumber")
            or fallback
        ).strip()

    @staticmethod
    def part_manufacturer(part: dict) -> str:
        return str(part.get("manufacturer") or part.get("Manufacturer") or "")

    @staticmethod
    def part_manufacturer_reference(part: dict) -> str:
        return str(
            part.get("manufacturer_part_number")
            or part.get("ManufacturerPartNumber")
            or ""
        )

    @staticmethod
    def part_description(part: dict) -> str:
        return str(part.get("description") or part.get("Description") or "")

    def list_components(self, sheet) -> list:
        """Lista componentes (ignora linhas vazias)."""
        items = []
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            if self.row_is_empty(row):
                continue
            items.append(self.row_to_dict(row))
        return items

    def row_to_dict(self, row) -> dict:
        location = ""
        if len(row) >= COMP_COL_LOCATION:
            location = str(row[COMP_COL_LOCATION - 1].value or "").strip()
        return {
            "mouser": row[1].value or "",
            "manufacturer": row[2].value or "",
            "manufacturer_ref": row[3].value or "",
            "description": row[5].value or "",
            "stock": row[6].value if row[6].value is not None else 0,
            "location": location,
            "locations": self.parse_component_locations(location),
        }

    def component_exists(
        self, sheet, mouser_ref: str, manufacturer_ref: str = ""
    ) -> bool:
        return self.find_component_any(sheet, mouser_ref, manufacturer_ref) is not None

    def get_history_rows(
        self, workbook, component_only: bool = False, mouser_ref: str = ""
    ) -> list:
        history = self.get_history_sheet(workbook)
        rows = list(history.iter_rows(min_row=2, values_only=True))
        if component_only and mouser_ref:
            rows = [
                row for row in rows if str(row[2]).strip() == mouser_ref.strip()
            ]
        return list(reversed(rows[-20:]))

    def get_component_recent_rows(self, workbook, *, limit: int = 20) -> list:
        """Last N Components sheet rows (newest last in Excel → reversed)."""
        sheet = self.get_components_sheet(workbook)
        rows: list = []
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            if self.row_is_empty(row):
                continue
            rows.append(row)
        if limit > 0:
            rows = rows[-limit:]
        return list(reversed(rows))

    def next_component_id(self, sheet) -> int:
        max_id = 0
        for row in sheet.iter_rows(min_row=2, min_col=1, max_col=1):
            try:
                max_id = max(max_id, int(row[0].value or 0))
            except (TypeError, ValueError):
                pass
        return max_id + 1

    def add_component_row(
        self,
        sheet,
        mouser_ref: str,
        manufacturer: str = "",
        manufacturer_ref: str = "",
        description: str = "",
        stock: int = 0,
        location: str = "",
    ) -> None:
        sheet.append([
            self.next_component_id(sheet),
            mouser_ref,
            manufacturer,
            manufacturer_ref,
            "",
            description,
            stock,
            str(location or "").strip(),
        ])

    def _find_by_manufacturer_pair(
        self, sheet, manufacturer: str, manufacturer_ref: str
    ):
        manufacturer_n = self.normalize_ref(manufacturer)
        manufacturer_ref_n = self.normalize_ref(manufacturer_ref)
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            if self.row_is_empty(row):
                continue
            row_manufacturer = self.normalize_ref(row[2].value)
            row_manufacturer_ref = self.normalize_ref(row[3].value)
            if (
                row_manufacturer == manufacturer_n
                and row_manufacturer_ref == manufacturer_ref_n
            ):
                return row
        return None

    def add_manual_component(
        self,
        user: str,
        supplier_reference: str = "",
        manufacturer: str = "",
        manufacturer_reference: str = "",
        description: str = "",
        initial_stock: int = 0,
        location: str = "",
    ) -> tuple[bool, str]:
        supplier_reference = str(supplier_reference).strip()
        manufacturer = str(manufacturer).strip()
        manufacturer_reference = str(manufacturer_reference).strip()
        description = str(description).strip()
        location = str(location).strip()
        if location:
            location = self.format_component_locations(
                self.parse_component_locations(location)
            )

        if initial_stock < 0:
            return False, "Initial stock cannot be negative."
        if not supplier_reference and not (manufacturer and manufacturer_reference):
            return (
                False,
                "Provide Supplier Reference OR both Manufacturer and Manufacturer Reference.",
            )

        workbook = self.get_workbook()
        sheet = self.get_components_sheet(workbook)
        history = self.get_history_sheet(workbook)

        if supplier_reference:
            if self.find_component_by_supplier_ref(sheet, supplier_reference):
                return False, "A component with this supplier reference already exists."
        elif self._find_by_manufacturer_pair(sheet, manufacturer, manufacturer_reference):
            return (
                False,
                "A component with this Manufacturer + Manufacturer Reference already exists.",
            )

        self.add_component_row(
            sheet,
            mouser_ref=str(supplier_reference),
            manufacturer=manufacturer,
            manufacturer_ref=manufacturer_reference,
            description=description,
            stock=initial_stock,
            location=location,
        )

        if initial_stock > 0:
            history_ref = supplier_reference or manufacturer_reference or "MANUAL"
            self.add_history(history, user, history_ref, "IN", initial_stock, initial_stock)

        if not self.save_workbook(workbook):
            return False, "Close stock.xlsx in Excel before saving."

        return True, "Manual component added successfully."

    def update_component(
        self,
        row,
        supplier_reference: str = "",
        manufacturer: str = "",
        manufacturer_reference: str = "",
        description: str = "",
        location: str = "",
    ) -> tuple[bool, str]:
        supplier_reference = str(supplier_reference).strip()
        manufacturer = str(manufacturer).strip()
        manufacturer_reference = str(manufacturer_reference).strip()
        description = str(description).strip()
        location = str(location).strip()
        if location:
            location = self.format_component_locations(
                self.parse_component_locations(location)
            )

        if not supplier_reference and not (manufacturer and manufacturer_reference):
            return (
                False,
                "Provide Supplier Reference OR both Manufacturer and Manufacturer Reference.",
            )

        workbook = self.get_workbook()
        sheet = self.get_components_sheet(workbook)
        target_row = row[0].row

        for other in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            if self.row_is_empty(other) or other[0].row == target_row:
                continue
            if supplier_reference and self.normalize_ref(
                other[1].value
            ) == self.normalize_ref(supplier_reference):
                return False, "Another component already uses this supplier reference."
            if (
                manufacturer
                and manufacturer_reference
                and self.normalize_ref(other[2].value)
                == self.normalize_ref(manufacturer)
                and self.normalize_ref(other[3].value)
                == self.normalize_ref(manufacturer_reference)
            ):
                return (
                    False,
                    "Another component already uses this Manufacturer + Manufacturer Reference.",
                )

        row[1].value = str(supplier_reference)
        row[2].value = manufacturer
        row[3].value = manufacturer_reference
        row[5].value = description
        sheet.cell(row=target_row, column=COMP_COL_LOCATION, value=location)

        if not self.save_workbook(workbook):
            return False, "Close stock.xlsx in Excel before saving."

        return True, "Component updated successfully."

    @staticmethod
    def parse_component_locations(value) -> list[str]:
        text = str(value or "").strip()
        if not text:
            return []
        if COMPONENT_LOCATION_SEPARATOR in text:
            return [
                part.strip()
                for part in text.split(COMPONENT_LOCATION_SEPARATOR)
                if part.strip()
            ]
        return [text]

    @staticmethod
    def format_component_locations(locations: list[str]) -> str:
        names: list[str] = []
        seen: set[str] = set()
        for name in locations:
            clean = str(name or "").strip()
            if not clean:
                continue
            key = clean.casefold()
            if key in seen:
                continue
            seen.add(key)
            names.append(clean)
        return COMPONENT_LOCATION_SEPARATOR.join(names)

    def set_component_location(
        self, row, location: str | list[str]
    ) -> tuple[bool, str]:
        if isinstance(location, list):
            value = self.format_component_locations(location)
        else:
            value = self.format_component_locations(
                self.parse_component_locations(location)
            )
        workbook = self.get_workbook()
        sheet = self.get_components_sheet(workbook)
        sheet.cell(row=row[0].row, column=COMP_COL_LOCATION, value=value)
        if not self.save_workbook(workbook):
            return False, "Close stock.xlsx in Excel before saving."
        count = len(self.parse_component_locations(value))
        if count == 0:
            return True, "Location cleared."
        if count == 1:
            return True, "Location updated."
        return True, f"{count} locations updated."

    # ------------------------------------------------------------------
    # Generic (resistors / capacitors — one row per spec)
    # ------------------------------------------------------------------
    @staticmethod
    def normalize_massive_type(text) -> str:
        value = str(text or "").strip().upper()
        if value in ("RESISTOR", "RES", "R"):
            return "R"
        if value in ("CAPACITOR", "CAP", "C"):
            return "C"
        return value

    @staticmethod
    def build_massive_name(
        part_type: str,
        value: str,
        tolerance: str,
        package: str,
        *,
        dielectric: str = "",
        voltage: str = "",
    ) -> str:
        kind = "Resistor" if part_type == "R" else "Capacitor" if part_type == "C" else "Part"
        parts = [str(value or "").strip(), str(tolerance or "").strip(), str(package or "").strip()]
        if part_type == "C":
            for extra in (str(dielectric or "").strip(), str(voltage or "").strip()):
                if extra:
                    parts.append(extra)
        label = " ".join(part for part in parts if part)
        return f"{kind} {label}".strip()

    @staticmethod
    def massive_identity_key(
        part_type: str,
        value: str,
        tolerance: str,
        package: str,
        dielectric: str = "",
    ) -> str:
        parts = (
            StockTracker.normalize_massive_type(part_type),
            StockTracker.normalize_ref(value),
            StockTracker.normalize_ref(tolerance),
            StockTracker.normalize_ref(package),
        )
        if StockTracker.normalize_massive_type(part_type) == "C":
            parts = (*parts, StockTracker.normalize_ref(dielectric))
        return "|".join(parts)

    def _massive_row_is_header(self, row) -> bool:
        return str(row[0].value or "").strip().upper() == "ID"

    def massive_row_to_dict(self, row) -> dict:
        try:
            stock = int(row[6].value or 0)
        except (TypeError, ValueError):
            stock = 0
        return {
            "id": row[0].value or "",
            "part_type": self.normalize_massive_type(row[1].value),
            "value": str(row[2].value or "").strip(),
            "tolerance": str(row[3].value or "").strip(),
            "package": str(row[4].value or "").strip(),
            "name": str(row[5].value or "").strip(),
            "stock": stock,
            "supplier_reference": str(row[7].value or "").strip(),
            "dielectric": str(row[8].value or "").strip(),
            "voltage": str(row[9].value or "").strip(),
            "notes": str(row[10].value or "").strip(),
            "location": (
                str(row[MASSIVE_COL_LOCATION - 1].value or "").strip()
                if len(row) >= MASSIVE_COL_LOCATION
                else ""
            ),
        }

    def next_massive_id(self, sheet) -> int:
        max_id = 0
        for row in sheet.iter_rows(min_row=2, min_col=1, max_col=1):
            try:
                max_id = max(max_id, int(row[0].value or 0))
            except (TypeError, ValueError):
                pass
        return max_id + 1

    def find_massive_by_identity(
        self,
        sheet,
        part_type: str,
        value: str,
        tolerance: str,
        package: str,
        dielectric: str = "",
    ):
        target = self.massive_identity_key(
            part_type, value, tolerance, package, dielectric=dielectric
        )
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            if self.row_is_empty(row) or self._massive_row_is_header(row):
                continue
            data = self.massive_row_to_dict(row)
            key = self.massive_identity_key(
                data["part_type"],
                data["value"],
                data["tolerance"],
                data["package"],
                dielectric=data["dielectric"],
            )
            if key == target:
                return row
        return None

    def find_massive_by_supplier_ref(self, sheet, supplier_reference: str):
        """Exact match on supplier reference column — for scan/barcode."""
        ref_n = self.normalize_ref(supplier_reference)
        if not ref_n:
            return None
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            if self.row_is_empty(row) or self._massive_row_is_header(row):
                continue
            data = self.massive_row_to_dict(row)
            if self.normalize_ref(data["supplier_reference"]) == ref_n:
                return row
        return None

    def find_massive(self, sheet, part_number: str):
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            if self.row_is_empty(row) or self._massive_row_is_header(row):
                continue
            data = self.massive_row_to_dict(row)
            if self.refs_match(part_number, data["supplier_reference"]):
                return row
        return None

    def find_massive_any(self, sheet, *queries: str):
        seen: set[str] = set()
        for query in queries:
            key = self.normalize_ref(query)
            if not key or key in seen:
                continue
            seen.add(key)
            row = self.find_massive_by_supplier_ref(sheet, query)
            if row:
                return row
        for query in queries:
            key = self.normalize_ref(query)
            if not key or key in seen:
                continue
            seen.add(key)
            row = self.find_massive(sheet, query)
            if row:
                return row
        return None

    @staticmethod
    def normalize_passive_value_token(text: str) -> str:
        value = (
            str(text or "")
            .strip()
            .upper()
            .replace("µ", "U")
            .replace("Ω", "OHM")
            .replace(" ", "")
        )
        if value.endswith("OHMS"):
            value = value[:-1]
        return value

    def passive_value_matches(self, query: str, stored_value: str) -> bool:
        """Match resistance/capacitance values (10k, 4K7, 100nF, 0, …)."""
        q = self.normalize_passive_value_token(query)
        stored = str(stored_value or "").strip()
        if not q or not stored:
            return False
        v = self.normalize_passive_value_token(stored)
        if q == v:
            return True
        if len(q) >= 2 and (q in v or v in q):
            return True
        if not q.endswith("F") and v.startswith(q) and "F" in v:
            return True
        return self.normalize_ref(q) in self.normalize_ref(stored)

    def search_massive_by_value(self, sheet, query: str) -> list:
        """Generic rows whose Value column matches the query (R/C value search)."""
        results = []
        q = str(query or "").strip()
        if not q:
            return results
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            if self.row_is_empty(row) or self._massive_row_is_header(row):
                continue
            data = self.massive_row_to_dict(row)
            if self.passive_value_matches(q, data["value"]):
                results.append(row)
        return results

    def excel_autocomplete_passive_values(self, sheet) -> list[str]:
        """Distinct passive values for search autocomplete."""
        seen: set[str] = set()
        values: list[str] = []
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            if self.row_is_empty(row) or self._massive_row_is_header(row):
                continue
            text = str(self.massive_row_to_dict(row).get("value") or "").strip()
            if len(text) < 1:
                continue
            key = self.normalize_passive_value_token(text)
            if key in seen:
                continue
            seen.add(key)
            values.append(text)
        return sorted(values, key=lambda item: self.normalize_passive_value_token(item))

    def excel_autocomplete_massive_supplier_refs(self, sheet) -> list[str]:
        """Supplier references on Generic sheet for barcode/scan autocomplete."""
        seen: set[str] = set()
        refs: list[str] = []
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            if self.row_is_empty(row) or self._massive_row_is_header(row):
                continue
            text = str(self.massive_row_to_dict(row).get("supplier_reference") or "").strip()
            if len(text) < 1:
                continue
            key = text.upper()
            if key in seen:
                continue
            seen.add(key)
            refs.append(text)
        return sorted(refs, key=lambda x: x.upper())

    def search_massive_all(self, sheet, query: str) -> list:
        """All Generic rows matching value, name, package, tolerance or supplier ref."""
        results = []
        q = self.normalize_ref(query)
        if not q:
            return results
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            if self.row_is_empty(row) or self._massive_row_is_header(row):
                continue
            data = self.massive_row_to_dict(row)
            haystack = " ".join(
                (
                    data["part_type"],
                    data["value"],
                    data["tolerance"],
                    data["package"],
                    data["name"],
                    data["supplier_reference"],
                    data["dielectric"],
                    data["voltage"],
                    data["notes"],
                    data["location"],
                )
            )
            if q in self.normalize_ref(haystack):
                results.append(row)
        return results

    def get_massive_rows(self, workbook, *, query: str = "") -> list:
        sheet = self.get_massive_sheet(workbook)
        matches = None
        q = str(query or "").strip()
        if q:
            matches = set(id(row) for row in self.search_massive_all(sheet, q))
        rows: list[tuple] = []
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            if self.row_is_empty(row) or self._massive_row_is_header(row):
                continue
            if matches is not None and id(row) not in matches:
                continue
            data = self.massive_row_to_dict(row)
            rows.append((
                data["id"],
                data["part_type"],
                data["value"],
                data["tolerance"],
                data["package"],
                data["name"],
                data["stock"],
            ))
        return list(reversed(rows[-20:]))

    def get_massive_recent_rows(self, workbook, *, limit: int = 20) -> list:
        """Last N Generic sheet rows (newest last in Excel → reversed)."""
        sheet = self.get_massive_sheet(workbook)
        rows: list = []
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            if self.row_is_empty(row) or self._massive_row_is_header(row):
                continue
            rows.append(row)
        if limit > 0:
            rows = rows[-limit:]
        return list(reversed(rows))

    def add_massive_item(
        self,
        user: str,
        part_type: str,
        value: str,
        tolerance: str,
        package: str,
        *,
        name: str = "",
        initial_stock: int = 0,
        supplier_reference: str = "",
        dielectric: str = "",
        voltage: str = "",
        notes: str = "",
        location: str = "",
    ) -> tuple[bool, str]:
        part_type = self.normalize_massive_type(part_type)
        value = str(value).strip()
        tolerance = str(tolerance).strip()
        package = str(package).strip()
        supplier_reference = str(supplier_reference).strip()
        dielectric = str(dielectric).strip()
        voltage = str(voltage).strip()
        notes = str(notes).strip()
        location = str(location).strip()
        name = str(name).strip()

        if part_type not in _MASSIVE_TYPES:
            return False, "Type must be R (resistor) or C (capacitor)."
        if not value:
            return False, "Value is required (e.g. 10k, 100nF)."
        if not package:
            return False, "Package is required (e.g. 0603, 0805)."
        if initial_stock < 0:
            return False, "Initial stock cannot be negative."

        if not name:
            name = self.build_massive_name(
                part_type,
                value,
                tolerance,
                package,
                dielectric=dielectric,
                voltage=voltage,
            )

        workbook = self.get_workbook()
        sheet = self.get_massive_sheet(workbook)
        history = self.get_history_sheet(workbook)

        if self.find_massive_by_identity(
            sheet, part_type, value, tolerance, package, dielectric=dielectric
        ):
            return False, "A passive item with this spec already exists."

        sheet.append([
            self.next_massive_id(sheet),
            part_type,
            value,
            tolerance,
            package,
            name,
            initial_stock,
            supplier_reference,
            dielectric,
            voltage,
            notes,
            location,
        ])

        if initial_stock > 0:
            self.add_history(history, user, name, "IN", initial_stock, initial_stock)

        if not self.save_workbook(workbook):
            return False, "Close stock.xlsx in Excel before saving."
        return True, "Passive item added successfully."

    def update_massive_item(
        self,
        row,
        part_type: str,
        value: str,
        tolerance: str,
        package: str,
        *,
        name: str = "",
        supplier_reference: str = "",
        dielectric: str = "",
        voltage: str = "",
        notes: str = "",
        location: str = "",
    ) -> tuple[bool, str]:
        part_type = self.normalize_massive_type(part_type)
        value = str(value).strip()
        tolerance = str(tolerance).strip()
        package = str(package).strip()
        supplier_reference = str(supplier_reference).strip()
        dielectric = str(dielectric).strip()
        voltage = str(voltage).strip()
        notes = str(notes).strip()
        location = str(location).strip()
        name = str(name).strip()

        if part_type not in _MASSIVE_TYPES:
            return False, "Type must be R (resistor) or C (capacitor)."
        if not value or not package:
            return False, "Value and Package are required."

        if not name:
            name = self.build_massive_name(
                part_type,
                value,
                tolerance,
                package,
                dielectric=dielectric,
                voltage=voltage,
            )

        workbook = self.get_workbook()
        sheet = self.get_massive_sheet(workbook)
        target_row = row[0].row

        duplicate = self.find_massive_by_identity(
            sheet, part_type, value, tolerance, package, dielectric=dielectric
        )
        if duplicate is not None and duplicate[0].row != target_row:
            return False, "Another passive item already uses this spec."

        sheet.cell(row=target_row, column=MASSIVE_COL_TYPE, value=part_type)
        sheet.cell(row=target_row, column=MASSIVE_COL_VALUE, value=value)
        sheet.cell(row=target_row, column=MASSIVE_COL_TOLERANCE, value=tolerance)
        sheet.cell(row=target_row, column=MASSIVE_COL_PACKAGE, value=package)
        sheet.cell(row=target_row, column=MASSIVE_COL_NAME, value=name)
        sheet.cell(row=target_row, column=MASSIVE_COL_SUPPLIER_REF, value=supplier_reference)
        sheet.cell(row=target_row, column=MASSIVE_COL_DIELECTRIC, value=dielectric)
        sheet.cell(row=target_row, column=MASSIVE_COL_VOLTAGE, value=voltage)
        sheet.cell(row=target_row, column=MASSIVE_COL_NOTES, value=notes)
        sheet.cell(row=target_row, column=MASSIVE_COL_LOCATION, value=location)

        if not self.save_workbook(workbook):
            return False, "Close stock.xlsx in Excel before saving."
        return True, "Passive item updated successfully."

    def update_massive_stock(
        self,
        user: str,
        row,
        quantity: int,
        movement: str,
    ) -> tuple[bool, str]:
        if quantity <= 0:
            return False, "Quantity must be greater than 0."

        workbook = self.get_workbook()
        sheet = self.get_massive_sheet(workbook)
        history = self.get_history_sheet(workbook)
        row_idx = row[0].row

        fresh_row = None
        for candidate in sheet.iter_rows(min_row=row_idx, max_row=row_idx):
            if not self.row_is_empty(candidate) and not self._massive_row_is_header(candidate):
                fresh_row = candidate
                break
        if fresh_row is None:
            return False, "Passive item not found."

        data = self.massive_row_to_dict(fresh_row)
        current = int(data["stock"])
        if movement == "IN":
            new_stock = current + quantity
        elif movement == "OUT":
            if current < quantity:
                return False, "Insufficient stock."
            new_stock = current - quantity
        else:
            return False, "Invalid movement."

        sheet.cell(row=row_idx, column=MASSIVE_COL_STOCK, value=new_stock)
        history_ref = data["name"] or data["value"] or f"GENERIC-{data['id']}"
        self.add_history(history, user, history_ref, movement, quantity, new_stock)

        if not self.save_workbook(workbook):
            return False, "Close stock.xlsx in Excel before saving."
        return True, "Stock updated and history saved."

    def get_known_locations(self, *, limit: int = 50) -> list[str]:
        """Distinct locations already used, most frequent first."""
        from collections import Counter

        workbook = self.get_workbook()
        counts: Counter[str] = Counter()

        components = self.get_components_sheet(workbook)
        for row in components.iter_rows(min_row=2, max_row=components.max_row):
            if self.row_is_empty(row):
                continue
            if len(row) >= COMP_COL_LOCATION:
                loc = str(row[COMP_COL_LOCATION - 1].value or "").strip()
                for part in self.parse_component_locations(loc):
                    counts[part] += 1

        generic = self.get_massive_sheet(workbook)
        for row in generic.iter_rows(min_row=2, max_row=generic.max_row):
            if self.row_is_empty(row) or self._massive_row_is_header(row):
                continue
            if len(row) >= MASSIVE_COL_LOCATION:
                loc = str(row[MASSIVE_COL_LOCATION - 1].value or "").strip()
                if loc:
                    counts[loc] += 1

        equipments = self.get_equipments_sheet(workbook)
        for row in equipments.iter_rows(min_row=2, max_row=equipments.max_row):
            if self.row_is_empty(row):
                continue
            loc = self._equipment_cell_str(
                equipments, row[0].row, EQ_COL_LOCATION
            ).strip()
            if loc:
                counts[loc] += 1

        return [loc for loc, _ in counts.most_common(limit)]

    # ------------------------------------------------------------------
    # Statistics
    # ------------------------------------------------------------------
    LOW_STOCK_THRESHOLD = 10  # legacy default; use get_low_stock_threshold()

    @property
    def low_stock_threshold(self) -> int:
        return get_low_stock_threshold()

    def get_weekly_movement_stats(self, weeks: int = 8) -> list[dict]:
        """IN/OUT totals grouped by ISO week (most recent last)."""
        from collections import defaultdict

        workbook = self.get_workbook()
        history = self.get_history_sheet(workbook)
        buckets: dict[str, dict[str, int]] = defaultdict(lambda: {"in": 0, "out": 0})

        for row in history.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 5:
                continue
            date_text = str(row[0] or "").strip()
            movement = str(row[3] or "").strip().upper()
            try:
                qty = int(row[4] or 0)
            except (TypeError, ValueError):
                continue
            if movement not in {"IN", "OUT"} or qty <= 0:
                continue
            try:
                dt = datetime.strptime(date_text[:10], "%Y-%m-%d")
            except ValueError:
                continue
            iso = dt.isocalendar()
            label = f"{iso.year}-W{iso.week:02d}"
            key = "in" if movement == "IN" else "out"
            buckets[label][key] += qty

        labels = sorted(buckets.keys())
        if weeks > 0:
            labels = labels[-weeks:]
        return [
            {"label": label, "in": buckets[label]["in"], "out": buckets[label]["out"]}
            for label in labels
        ]

    @staticmethod
    def _expiry_sort_key(expiry_str: str) -> tuple[int, date]:
        text = str(expiry_str or "").strip()
        if not text:
            return (1, date.max)
        normalized = StockTracker.normalize_date(text)
        try:
            return (0, datetime.strptime(normalized, "%Y-%m-%d").date())
        except ValueError:
            return (1, date.max)

    def get_equipment_expiration_stats(self) -> list[dict]:
        """Equipments sorted by calibration expiration (soonest first)."""
        workbook = self.get_workbook()
        sheet = self.get_equipments_sheet(workbook)
        today = date.today()
        items: list[dict] = []
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            if self.row_is_empty(row):
                continue
            data = self.equipment_row_to_dict(row)
            sort_key = self._expiry_sort_key(data.get("calibration_expiration", ""))
            days_left = None
            if sort_key[0] == 0:
                days_left = (sort_key[1] - today).days
            items.append({**data, "_sort": sort_key, "days_left": days_left})
        items.sort(key=lambda item: item["_sort"])
        for item in items:
            item.pop("_sort", None)
        return items

    def get_low_stock_components(self, threshold: int | None = None) -> list[dict]:
        """Components at or below stock threshold, lowest stock first."""
        limit = self.low_stock_threshold if threshold is None else threshold
        workbook = self.get_workbook()
        sheet = self.get_components_sheet(workbook)
        items: list[dict] = []
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            if self.row_is_empty(row):
                continue
            data = self.row_to_dict(row)
            try:
                stock = int(data.get("stock") or 0)
            except (TypeError, ValueError):
                stock = 0
            if stock > limit:
                continue
            label = (
                str(data.get("description") or "").strip()
                or str(data.get("mouser") or "").strip()
                or str(data.get("manufacturer_ref") or "").strip()
            )
            items.append({**data, "stock": stock, "label": label})
        items.sort(key=lambda item: (item["stock"], item["label"].lower()))
        return items

    def get_low_stock_massive(self, threshold: int | None = None) -> list[dict]:
        """Generic items at or below stock threshold, lowest stock first."""
        limit = self.low_stock_threshold if threshold is None else threshold
        workbook = self.get_workbook()
        sheet = self.get_massive_sheet(workbook)
        items: list[dict] = []
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            if self.row_is_empty(row) or self._massive_row_is_header(row):
                continue
            data = self.massive_row_to_dict(row)
            if int(data.get("stock") or 0) > limit:
                continue
            items.append(data)
        items.sort(
            key=lambda item: (
                int(item.get("stock") or 0),
                str(item.get("name") or "").lower(),
            )
        )
        return items

    def get_items_without_location(self) -> list[dict]:
        """Inventory rows with stock > 0 and no Location (Components + Generic)."""
        workbook = self.get_workbook()
        items: list[dict] = []

        components = self.get_components_sheet(workbook)
        for row in components.iter_rows(min_row=2, max_row=components.max_row):
            if self.row_is_empty(row):
                continue
            stock = self._stock_cell_int(row[6].value)
            if stock is None or stock <= 0:
                continue
            data = self.row_to_dict(row)
            if data.get("locations"):
                continue
            items.append(
                {
                    "sheet": "Components",
                    "id": row[0].value,
                    "reference": str(data.get("mouser") or ""),
                    "description": str(data.get("description") or ""),
                    "stock": stock,
                }
            )

        generic = self.get_massive_sheet(workbook)
        for row in generic.iter_rows(min_row=2, max_row=generic.max_row):
            if self.row_is_empty(row) or self._massive_row_is_header(row):
                continue
            stock = self._stock_cell_int(row[6].value)
            if stock is None or stock <= 0:
                continue
            data = self.massive_row_to_dict(row)
            if str(data.get("location") or "").strip():
                continue
            ref = str(data.get("supplier_reference") or data.get("value") or "")
            desc = (
                f"{data.get('part_type', '')} {data.get('value', '')} "
                f"{data.get('tolerance', '')} {data.get('package', '')} — "
                f"{data.get('name', '')}"
            ).strip()
            items.append(
                {
                    "sheet": "Generic",
                    "id": data.get("id"),
                    "reference": ref,
                    "description": desc,
                    "stock": stock,
                }
            )

        items.sort(key=lambda item: (item["sheet"], str(item.get("reference") or "").lower()))
        return items

    def set_massive_location(self, row, location: str) -> tuple[bool, str]:
        location = str(location or "").strip()
        workbook = self.get_workbook()
        sheet = self.get_massive_sheet(workbook)
        sheet.cell(row=row[0].row, column=MASSIVE_COL_LOCATION, value=location)
        if not self.save_workbook(workbook):
            return False, "Close stock.xlsx in Excel before saving."
        return True, "Location updated."

    def get_loaned_equipments(self) -> list[dict]:
        """Equipments currently marked as on loan."""
        workbook = self.get_workbook()
        sheet = self.get_equipments_sheet(workbook)
        items: list[dict] = []
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            if self.row_is_empty(row):
                continue
            data = self.equipment_row_to_dict(row)
            if data.get("loaned"):
                items.append(data)
        items.sort(
            key=lambda item: str(item.get("loan_since") or "").lower(),
            reverse=True,
        )
        return items

    UNASSIGNED_LOCATION = "(no location)"

    def get_location_statistics(self, threshold: int | None = None) -> list[dict]:
        """Per-location counts: components, passive, equipments, low stock, stock units."""
        from collections import defaultdict

        limit = self.low_stock_threshold if threshold is None else threshold
        workbook = self.get_workbook()

        buckets: dict[str, dict[str, int]] = defaultdict(
            lambda: {
                "components": 0,
                "passive": 0,
                "equipments": 0,
                "low_stock": 0,
                "stock_units": 0,
            }
        )

        def _touch(location: str) -> str:
            text = str(location or "").strip()
            return text or self.UNASSIGNED_LOCATION

        def _record(
            location: str,
            *,
            components: int = 0,
            passive: int = 0,
            equipments: int = 0,
            stock: int = 0,
            is_low: bool = False,
        ) -> None:
            key = _touch(location)
            bucket = buckets[key]
            bucket["components"] += components
            bucket["passive"] += passive
            bucket["equipments"] += equipments
            if stock:
                bucket["stock_units"] += stock
            if is_low:
                bucket["low_stock"] += 1

        components_sheet = self.get_components_sheet(workbook)
        for row in components_sheet.iter_rows(min_row=2, max_row=components_sheet.max_row):
            if self.row_is_empty(row):
                continue
            data = self.row_to_dict(row)
            try:
                stock = int(data.get("stock") or 0)
            except (TypeError, ValueError):
                stock = 0
            is_low = stock <= limit
            locations = data.get("locations") or self.parse_component_locations(
                data.get("location", "")
            )
            if not locations:
                _record("", components=1, stock=stock, is_low=is_low)
                continue
            for loc in locations:
                _record(loc, components=1, stock=stock, is_low=is_low)

        massive_sheet = self.get_massive_sheet(workbook)
        for row in massive_sheet.iter_rows(min_row=2, max_row=massive_sheet.max_row):
            if self.row_is_empty(row) or self._massive_row_is_header(row):
                continue
            data = self.massive_row_to_dict(row)
            try:
                stock = int(data.get("stock") or 0)
            except (TypeError, ValueError):
                stock = 0
            is_low = stock <= limit
            loc = str(data.get("location") or "").strip()
            _record(loc, passive=1, stock=stock, is_low=is_low)

        equipments_sheet = self.get_equipments_sheet(workbook)
        for row in equipments_sheet.iter_rows(min_row=2, max_row=equipments_sheet.max_row):
            if self.row_is_empty(row):
                continue
            data = self.equipment_row_to_dict(row)
            loc = str(data.get("location") or "").strip()
            _record(loc, equipments=1)

        items: list[dict] = []
        for location, counts in buckets.items():
            total_items = (
                counts["components"] + counts["passive"] + counts["equipments"]
            )
            items.append(
                {
                    "location": location,
                    "components": counts["components"],
                    "passive": counts["passive"],
                    "equipments": counts["equipments"],
                    "low_stock": counts["low_stock"],
                    "stock_units": counts["stock_units"],
                    "total_items": total_items,
                }
            )

        items.sort(
            key=lambda item: (
                item["location"] == self.UNASSIGNED_LOCATION,
                -item["total_items"],
                item["location"].lower(),
            )
        )
        return items

    def get_inventory_distribution(self) -> list[dict]:
        """Counts and stock totals per inventory category (for charts)."""
        workbook = self.get_workbook()
        components_sheet = self.get_components_sheet(workbook)
        massive_sheet = self.get_massive_sheet(workbook)
        equipments_sheet = self.get_equipments_sheet(workbook)

        components_count = 0
        components_stock = 0
        for row in components_sheet.iter_rows(min_row=2, max_row=components_sheet.max_row):
            if self.row_is_empty(row):
                continue
            components_count += 1
            try:
                components_stock += int(row[6].value or 0)
            except (TypeError, ValueError):
                pass

        massive_count = 0
        massive_stock = 0
        for row in massive_sheet.iter_rows(min_row=2, max_row=massive_sheet.max_row):
            if self.row_is_empty(row) or self._massive_row_is_header(row):
                continue
            massive_count += 1
            try:
                massive_stock += int(row[6].value or 0)
            except (TypeError, ValueError):
                pass

        equipments_count = 0
        for row in equipments_sheet.iter_rows(min_row=2, max_row=equipments_sheet.max_row):
            if self.row_is_empty(row):
                continue
            equipments_count += 1

        return [
            {
                "key": "components",
                "label": "Components",
                "count": components_count,
                "stock": components_stock,
            },
            {
                "key": "generic",
                "label": "Passive (R/C)",
                "count": massive_count,
                "stock": massive_stock,
            },
            {
                "key": "equipments",
                "label": "Equipments",
                "count": equipments_count,
                "stock": equipments_count,
            },
        ]

    # ------------------------------------------------------------------
    # Distribuidores (Mouser, TME, ...)
    # ------------------------------------------------------------------
    def search_supplier(
        self, supplier: SupplierId, part_number: str
    ) -> Optional[dict]:
        """Pesquisa num fornecedor; devolve dict compativel com codigo existente."""
        part = search_part(supplier, part_number, self._secrets)
        return dict(part) if part else None

    def search_mouser(self, part_number: str) -> Optional[dict]:
        return self.search_supplier("mouser", part_number)

    def search_digikey(self, part_number: str) -> Optional[dict]:
        return self.search_supplier("digikey", part_number)

    def search_rs(self, part_number: str) -> Optional[dict]:
        return self.search_supplier("rs", part_number)

    def search_tme(self, part_number: str) -> Optional[dict]:
        return self.search_supplier("tme", part_number)

    def configured_suppliers(self) -> list[str]:
        from .suppliers.credentials import configured_suppliers

        return configured_suppliers(self._secrets)

    def search_suppliers_order(self) -> list[SupplierId]:
        """Fornecedores configurados, por ordem de pesquisa (APIs estaveis primeiro)."""
        configured = set(self.configured_suppliers())
        return [s for s in _SUPPLIER_SEARCH_ORDER if s in configured]

    def search_any_supplier(
        self, part_number: str
    ) -> tuple[Optional[dict], Optional[SupplierId]]:
        """Pesquisa em cada distribuidor configurado ate encontrar resultado."""
        for supplier in self.search_suppliers_order():
            part = self.search_supplier(supplier, part_number)
            if part is not None:
                return part, supplier
        return None, None

    @staticmethod
    def _catalog_key_usable(key: str) -> bool:
        """Short text-only queries are Excel search terms, not catalog part numbers."""
        if len(key) >= 5:
            return True
        return bool(re.search(r"\d", key))

    def _catalog_part_matches_query(self, part_number: str, part: dict) -> bool:
        """Reject vague API hits (e.g. Mouser returning an unrelated first result)."""
        return self.refs_match(
            part_number,
            part.get("supplier_part_number"),
            part.get("MouserPartNumber"),
            part.get("manufacturer_part_number"),
            part.get("ManufacturerPartNumber"),
        )

    def _merge_catalog_part(self, base: dict, fresh: dict | None) -> dict:
        merged = dict(base)
        if fresh:
            for key, value in fresh.items():
                if value is not None and str(value).strip():
                    merged[key] = value
        return merged

    def _enrich_catalog_part(self, part: dict) -> dict:
        from .component_datasheet_urls import resolve_datasheet_url

        merged = dict(part)
        datasheet = resolve_datasheet_url(merged)
        merged["datasheet_url"] = datasheet
        return merged

    def lookup_catalog_part(self, part_number: str) -> Optional[dict]:
        """Return distributor catalog fields (URLs, image) with session + disk cache."""
        key = self.normalize_ref(part_number)
        if not key:
            return None

        from .component_catalog_links import get_cached_links, store_links

        with _CATALOG_LOOKUP_LOCK:
            session = self._catalog_session_cache.get(key)
            if session is not None and self._catalog_key_usable(key):
                merged = self._merge_catalog_part(session, None)
                if not str(merged.get("datasheet_url", "")).strip():
                    merged = self._merge_catalog_part(
                        merged, self.search_mouser(part_number)
                    )
                enriched = self._enrich_catalog_part(merged)
                if self._catalog_part_matches_query(
                    part_number,
                    enriched,
                ) or str(enriched.get("product_url", "")).strip():
                    self._catalog_session_cache[key] = enriched
                    return enriched
                self._catalog_session_cache.pop(key, None)

            if self._catalog_key_usable(key):
                cached = get_cached_links(key)
                if cached is not None:
                    merged = dict(cached)
                    if not str(merged.get("datasheet_url", "")).strip():
                        merged = self._merge_catalog_part(
                            merged, self.search_mouser(part_number)
                        )
                    enriched = self._enrich_catalog_part(merged)
                    if str(enriched.get("product_url", "")).strip() or str(
                        enriched.get("datasheet_url", "")
                    ).strip():
                        self._catalog_session_cache[key] = enriched
                        if enriched.get("datasheet_url") != cached.get(
                            "datasheet_url"
                        ) or enriched.get("manufacturer_part_number") != cached.get(
                            "manufacturer_part_number"
                        ):
                            store_links(key, enriched)
                        return enriched

            part = self.search_mouser(part_number)
            if part is None:
                part, _supplier = self.search_any_supplier(part_number)
            if part is None:
                return None
            if not self._catalog_part_matches_query(part_number, part):
                return None

            merged = self._enrich_catalog_part(part)
            merged.update(store_links(key, merged))
            self._catalog_session_cache[key] = merged
            return merged

    def lookup_catalog_part_any(self, *part_numbers: str) -> Optional[dict]:
        """Try several references until a distributor catalog match is found."""
        seen: set[str] = set()
        for part_number in part_numbers:
            key = self.normalize_ref(part_number)
            if not key or key in seen:
                continue
            seen.add(key)
            part = self.lookup_catalog_part(part_number)
            if part is not None:
                return part
        return None

    def lookup_catalog_image_url(self, part_number: str) -> str:
        """Return a product image URL from configured distributor APIs."""
        part = self.lookup_catalog_part_any(part_number)
        if part is None:
            return ""
        return str(part.get("image_url", "")).strip()

    def lookup_catalog_image_url_any(self, *part_numbers: str) -> str:
        part = self.lookup_catalog_part_any(*part_numbers)
        if part is None:
            return ""
        return str(part.get("image_url", "")).strip()

    def lookup_catalog_links(self, part_number: str) -> tuple[str, str]:
        """Return (product_page_url, datasheet_url) for a component reference."""
        part = self.lookup_catalog_part(part_number)
        if part is None:
            return "", ""
        return (
            str(part.get("product_url", "")).strip(),
            str(part.get("datasheet_url", "")).strip(),
        )

    def lookup_catalog_links_any(self, *part_numbers: str) -> tuple[str, str]:
        part = self.lookup_catalog_part_any(*part_numbers)
        if part is None:
            return "", ""
        return (
            str(part.get("product_url", "")).strip(),
            str(part.get("datasheet_url", "")).strip(),
        )

    def configured_supplier_labels(self) -> str:
        return ", ".join(supplier_label(s) for s in self.search_suppliers_order())

    # ------------------------------------------------------------------
    # Stock movements
    # ------------------------------------------------------------------
    def add_history(self, history, user, mouser_ref, movement, qty, stock_after):
        history.append([
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            user,
            mouser_ref,
            movement,
            qty,
            stock_after,
        ])

    def update_stock(self, user: str, code: str, quantity: int, movement: str) -> bool:
        """
        movement: "IN" or "OUT"
        Returns True if saved successfully.
        """
        workbook = self.get_workbook()
        sheet = self.get_components_sheet(workbook)
        history = self.get_history_sheet(workbook)

        part_number = self.extract_part_number(code)
        row = self.find_component_any(sheet, part_number, code)
        if row is None:
            print("Componente nao encontrado.")
            return False

        current = int(row[6].value or 0)
        if movement == "IN":
            new_stock = current + quantity
        else:
            if current < quantity:
                print("Stock insuficiente.")
                return False
            new_stock = current - quantity

        row[6].value = new_stock
        self.add_history(
            history, user, row[1].value, movement, quantity, new_stock
        )
        return self.save_workbook(workbook)

    def add_from_supplier_and_stock_in(
        self,
        user: str,
        code: str,
        quantity: int,
        supplier: SupplierId | None = None,
    ) -> bool:
        """
        Full flow:
        1) If component exists in Excel -> add stock (IN)
        2) If not -> search distributor API(s) -> add row -> add stock (IN)
        """
        if quantity <= 0:
            print("Quantity must be greater than 0.")
            return False

        workbook = self.get_workbook()
        sheet = self.get_components_sheet(workbook)
        history = self.get_history_sheet(workbook)

        part_number = self.extract_part_number(code)
        row = self.find_component_any(sheet, part_number, code)

        if row is None:
            if supplier is not None:
                part = self.search_supplier(supplier, part_number)
            else:
                part, _found = self.search_any_supplier(part_number)
            if part is None:
                print("Not found in Excel or distributor catalog.")
                return False

            supplier_ref = self.part_supplier_reference(part, part_number)
            manufacturer_ref = self.part_manufacturer_reference(part)
            if self.find_component_any(sheet, supplier_ref, manufacturer_ref):
                row = self.find_component_any(
                    sheet, supplier_ref, manufacturer_ref, part_number, code
                )
            else:
                self.add_component_row(
                    sheet,
                    mouser_ref=supplier_ref,
                    manufacturer=self.part_manufacturer(part),
                    manufacturer_ref=manufacturer_ref,
                    description=self.part_description(part),
                    stock=0,
                )
                if not self.save_workbook(workbook):
                    return False
                print(f"New component added to Excel: {supplier_ref}")
                workbook = self.get_workbook()
                sheet = self.get_components_sheet(workbook)
                history = self.get_history_sheet(workbook)
                row = self.find_component_any(
                    sheet, supplier_ref, manufacturer_ref, part_number, code
                )

        if row is None:
            print("Could not locate component after save.")
            return False

        current = int(row[6].value or 0)
        new_stock = current + quantity
        row[6].value = new_stock
        self.add_history(
            history, user, row[1].value, "IN", quantity, new_stock
        )
        if self.save_workbook(workbook):
            print(f"Stock updated: {new_stock}")
            return True
        return False

    def add_from_mouser_and_stock_in(
        self, user: str, code: str, quantity: int
    ) -> bool:
        """Pesquisa em todos os distribuidores configurados (nome legado)."""
        return self.add_from_supplier_and_stock_in(user, code, quantity)

    # ------------------------------------------------------------------
    # Equipments (calibration tracking)
    # ------------------------------------------------------------------
    @staticmethod
    def _migrate_legacy_materials_sheet_name(workbook) -> None:
        """Rename legacy 'Materials' Excel sheet to 'Equipments'."""
        if (
            SHEET_MATERIALS_LEGACY in workbook.sheetnames
            and SHEET_EQUIPMENTS not in workbook.sheetnames
        ):
            workbook[SHEET_MATERIALS_LEGACY].title = SHEET_EQUIPMENTS

    def get_equipments_sheet(self, workbook):
        self._migrate_legacy_materials_sheet_name(workbook)
        if SHEET_EQUIPMENTS not in workbook.sheetnames:
            sheet = workbook.create_sheet(SHEET_EQUIPMENTS)
        else:
            sheet = workbook[SHEET_EQUIPMENTS]

        if sheet.max_row == 0 or (
            sheet.max_row == 1 and sheet["A1"].value is None
        ):
            headers = [
                "ID",
                "Supplier Reference",
                "Serial Number",
                "Name",
                "Description",
                "Calibration Date",
                "Calibration Expiration Date",
                "Datasheet",
                "Image",
                "Location",
                "Loaned",
                "Loaned To",
                "Loan Place",
                "Loan Since",
            ]
            if sheet.max_row == 0:
                sheet.append(headers)
            else:
                for col, value in enumerate(headers, start=1):
                    sheet.cell(row=1, column=col, value=value)
        else:
            self._migrate_equipments_sheet(sheet)
        return sheet

    @staticmethod
    def _migrate_equipments_sheet(sheet) -> None:
        """Upgrade legacy Equipments sheets to the current column layout."""
        if str(sheet["A1"].value or "").strip() != "ID":
            return
        b1 = str(sheet["B1"].value or "").strip()
        if b1 == "Description":
            sheet.insert_cols(2)
            sheet["B1"] = "Supplier Reference"
            b1 = "Supplier Reference"
        c1 = str(sheet["C1"].value or "").strip()
        if b1 == "Supplier Reference" and c1 == "Description":
            sheet.insert_cols(3)
            sheet["C1"] = "Serial Number"
        g1 = str(sheet.cell(row=1, column=EQ_COL_DATASHEET).value or "").strip()
        if sheet.max_column < EQ_COL_DATASHEET or not g1:
            sheet.cell(row=1, column=EQ_COL_DATASHEET, value="Datasheet")
        h1 = str(sheet.cell(row=1, column=EQ_COL_IMAGE).value or "").strip()
        if sheet.max_column < EQ_COL_IMAGE or not h1:
            sheet.cell(row=1, column=EQ_COL_IMAGE, value="Image")
        d1 = str(sheet.cell(row=1, column=EQ_COL_NAME).value or "").strip()
        if d1 == "Description":
            sheet.insert_cols(EQ_COL_NAME)
            sheet.cell(row=1, column=EQ_COL_NAME, value="Name")
        loan_headers = {
            EQ_COL_LOCATION: "Location",
            EQ_COL_LOANED: "Loaned",
            EQ_COL_LOANED_TO: "Loaned To",
            EQ_COL_LOAN_PLACE: "Loan Place",
            EQ_COL_LOAN_SINCE: "Loan Since",
        }
        for col, header in loan_headers.items():
            current = str(sheet.cell(row=1, column=col).value or "").strip()
            if sheet.max_column < col or not current:
                sheet.cell(row=1, column=col, value=header)

    def get_equipment_loans_sheet(self, workbook):
        if SHEET_EQUIPMENT_LOANS not in workbook.sheetnames:
            sheet = workbook.create_sheet(SHEET_EQUIPMENT_LOANS)
        else:
            sheet = workbook[SHEET_EQUIPMENT_LOANS]

        if sheet.max_row == 0 or (
            sheet.max_row == 1 and sheet["A1"].value is None
        ):
            headers = [
                "Date",
                "User",
                "Equipment ID",
                "Action",
                "Loaned To",
                "Place",
                "Home Location",
                "Notes",
            ]
            if sheet.max_row == 0:
                sheet.append(headers)
            else:
                for col, value in enumerate(headers, start=1):
                    sheet.cell(row=1, column=col, value=value)
        return sheet

    @staticmethod
    def _loan_datetime_now() -> str:
        return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    def _append_equipment_loan_record(
        self,
        workbook,
        *,
        user: str,
        equipment_id: str | int,
        action: str,
        loaned_to: str = "",
        place: str = "",
        home_location: str = "",
        notes: str = "",
    ) -> None:
        loans = self.get_equipment_loans_sheet(workbook)
        loans.append([
            self._loan_datetime_now(),
            str(user).strip(),
            str(equipment_id).strip(),
            str(action).strip().upper(),
            str(loaned_to).strip(),
            str(place).strip(),
            str(home_location).strip(),
            str(notes).strip(),
        ])

    @staticmethod
    def _equipment_cell_str(sheet, row_idx: int, column: int) -> str:
        if sheet is None:
            return ""
        value = sheet.cell(row=row_idx, column=column).value
        if value is None:
            return ""
        return str(value).strip()

    def _equipment_loan_fields_from_sheet(self, sheet, row_idx: int) -> dict:
        loaned_raw = self._equipment_cell_str(sheet, row_idx, EQ_COL_LOANED)
        loaned = loaned_raw.lower() in {"yes", "y", "sim", "1", "true"}
        return {
            "location": self._equipment_cell_str(sheet, row_idx, EQ_COL_LOCATION),
            "loaned": loaned,
            "loaned_to": self._equipment_cell_str(sheet, row_idx, EQ_COL_LOANED_TO),
            "loan_place": self._equipment_cell_str(sheet, row_idx, EQ_COL_LOAN_PLACE),
            "loan_since": self._equipment_cell_str(sheet, row_idx, EQ_COL_LOAN_SINCE),
        }

    @staticmethod
    def normalize_date(text) -> str:
        """Return YYYY-MM-DD or empty string."""
        if text is None:
            return ""
        value = str(text).strip()
        if not value:
            return ""
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(value, fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
        return value

    @staticmethod
    def validate_date(text: str) -> tuple[bool, str]:
        raw = str(text or "").strip()
        if not raw:
            return True, ""
        normalized = StockTracker.normalize_date(raw)
        try:
            datetime.strptime(normalized, "%Y-%m-%d")
            return True, normalized
        except ValueError:
            return False, ""

    def search_equipments_all(self, sheet, query: str) -> list:
        """All equipment rows matching supplier reference or description."""
        results = []
        q = self.normalize_ref(query)
        if not q:
            return results
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            if self.row_is_empty(row):
                continue
            supplier_ref = self.normalize_ref(row[1].value)
            serial = self.normalize_ref(row[2].value)
            name = self.normalize_ref(row[3].value)
            desc = self.normalize_ref(row[4].value)
            if (
                q in supplier_ref
                or supplier_ref in q
                or q in serial
                or serial in q
                or q in name
                or name in q
                or q in desc
                or desc in q
            ):
                results.append(row)
        return results

    def find_equipment_by_supplier_ref(self, sheet, supplier_ref: str):
        key = self.normalize_ref(supplier_ref)
        if not key:
            return None
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            if self.row_is_empty(row):
                continue
            if self.normalize_ref(row[1].value) == key:
                return row
        return None

    def equipment_row_to_dict(self, row) -> dict:
        datasheet = ""
        image = ""
        row_idx = row[0].row
        sheet = row[0].parent
        if sheet is not None:
            cell_val = sheet.cell(row=row_idx, column=EQ_COL_DATASHEET).value
            if cell_val is not None:
                datasheet = str(cell_val).strip()
            image_val = sheet.cell(row=row_idx, column=EQ_COL_IMAGE).value
            if image_val is not None:
                image = str(image_val).strip()
            name = sheet.cell(row=row_idx, column=EQ_COL_NAME).value or ""
            description = sheet.cell(row=row_idx, column=EQ_COL_DESCRIPTION).value or ""
            calibration_date = sheet.cell(row=row_idx, column=EQ_COL_CALIB_DATE).value
            calibration_expiration = sheet.cell(
                row=row_idx, column=EQ_COL_CALIB_EXPIRY
            ).value
            loan_fields = self._equipment_loan_fields_from_sheet(sheet, row_idx)
        else:
            name = row[3].value if len(row) > 3 else ""
            description = row[4].value if len(row) > 4 else ""
            calibration_date = row[5].value if len(row) > 5 else ""
            calibration_expiration = row[6].value if len(row) > 6 else ""
            if len(row) > 7 and row[7].value is not None:
                datasheet = str(row[7].value).strip()
            if len(row) > 8 and row[8].value is not None:
                image = str(row[8].value).strip()
            loan_fields = {
                "location": "",
                "loaned": False,
                "loaned_to": "",
                "loan_place": "",
                "loan_since": "",
            }
        return {
            "id": row[0].value or "",
            "supplier_reference": row[1].value or "",
            "serial_number": row[2].value or "",
            "name": str(name or "").strip(),
            "description": str(description or "").strip(),
            "calibration_date": self.normalize_date(calibration_date),
            "calibration_expiration": self.normalize_date(calibration_expiration),
            "datasheet": datasheet,
            "image": image,
            "images": self.parse_equipment_images(image),
            **loan_fields,
        }

    def next_equipment_id(self, sheet) -> int:
        max_id = 0
        for row in sheet.iter_rows(min_row=2, min_col=1, max_col=1):
            try:
                max_id = max(max_id, int(row[0].value or 0))
            except (TypeError, ValueError):
                pass
        return max_id + 1

    def add_equipment(
        self,
        description: str = "",
        supplier_reference: str = "",
        serial_number: str = "",
        name: str = "",
        calibration_date: str = "",
        calibration_expiration: str = "",
        datasheet: str = "",
    ) -> tuple[bool, str]:
        description = str(description).strip()
        name = str(name).strip()
        supplier_reference = str(supplier_reference).strip()
        serial_number = str(serial_number).strip()
        if not description and not name and not supplier_reference and not serial_number:
            return False, "Provide Name, Description, Supplier Reference or Serial Number."

        ok, calib = self.validate_date(calibration_date)
        if not ok:
            return False, "Invalid calibration date (use YYYY-MM-DD)."
        ok, expiry = self.validate_date(calibration_expiration)
        if not ok:
            return False, "Invalid expiration date (use YYYY-MM-DD)."

        workbook = self.get_workbook()
        sheet = self.get_equipments_sheet(workbook)
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            if self.row_is_empty(row):
                continue
            if supplier_reference and self.normalize_ref(
                row[1].value
            ) == self.normalize_ref(supplier_reference):
                return False, "An equipment with this supplier reference already exists."
            if serial_number and self.normalize_ref(
                row[2].value
            ) == self.normalize_ref(serial_number):
                return False, "An equipment with this serial number already exists."
            if description and self.normalize_ref(
                row[4].value
            ) == self.normalize_ref(description):
                return False, "An equipment with this description already exists."
            if name and self.normalize_ref(row[3].value) == self.normalize_ref(name):
                return False, "An equipment with this name already exists."

        sheet.append([
            self.next_equipment_id(sheet),
            supplier_reference,
            serial_number,
            name,
            description,
            calib,
            expiry,
            str(datasheet).strip(),
            "",
        ])
        if not self.save_workbook(workbook):
            return False, "Close stock.xlsx in Excel before saving."
        return True, "Equipment added successfully."

    def update_equipment(
        self,
        row,
        description: str = "",
        supplier_reference: str = "",
        serial_number: str = "",
        name: str = "",
        calibration_date: str = "",
        calibration_expiration: str = "",
        datasheet: str | None = None,
    ) -> tuple[bool, str]:
        description = str(description).strip()
        name = str(name).strip()
        supplier_reference = str(supplier_reference).strip()
        serial_number = str(serial_number).strip()
        if not description and not name and not supplier_reference and not serial_number:
            return False, "Provide Name, Description, Supplier Reference or Serial Number."

        ok, calib = self.validate_date(calibration_date)
        if not ok:
            return False, "Invalid calibration date (use YYYY-MM-DD)."
        ok, expiry = self.validate_date(calibration_expiration)
        if not ok:
            return False, "Invalid expiration date (use YYYY-MM-DD)."

        workbook = self.get_workbook()
        sheet = self.get_equipments_sheet(workbook)
        target_row = row[0].row
        ref_norm = self.normalize_ref(supplier_reference)
        serial_norm = self.normalize_ref(serial_number)
        name_norm = self.normalize_ref(name)
        desc_norm = self.normalize_ref(description)

        for other in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            if self.row_is_empty(other) or other[0].row == target_row:
                continue
            if supplier_reference and self.normalize_ref(other[1].value) == ref_norm:
                return False, "Another equipment already uses this supplier reference."
            if serial_number and self.normalize_ref(other[2].value) == serial_norm:
                return False, "Another equipment already uses this serial number."
            if name and self.normalize_ref(
                sheet.cell(row=other[0].row, column=EQ_COL_NAME).value
            ) == name_norm:
                return False, "Another equipment already uses this name."
            if description and self.normalize_ref(
                sheet.cell(row=other[0].row, column=EQ_COL_DESCRIPTION).value
            ) == desc_norm:
                return False, "Another equipment already uses this description."

        sheet.cell(row=target_row, column=EQ_COL_SUPPLIER_REF, value=supplier_reference)
        sheet.cell(row=target_row, column=EQ_COL_SERIAL, value=serial_number)
        sheet.cell(row=target_row, column=EQ_COL_NAME, value=name)
        sheet.cell(row=target_row, column=EQ_COL_DESCRIPTION, value=description)
        sheet.cell(row=target_row, column=EQ_COL_CALIB_DATE, value=calib)
        sheet.cell(row=target_row, column=EQ_COL_CALIB_EXPIRY, value=expiry)
        if datasheet is not None:
            sheet.cell(
                row=target_row,
                column=EQ_COL_DATASHEET,
                value=str(datasheet).strip(),
            )

        if not self.save_workbook(workbook):
            return False, "Close stock.xlsx in Excel before saving."
        return True, "Equipment updated successfully."

    def link_equipment_datasheet(self, row, filename: str) -> tuple[bool, str]:
        """Associate a support-document filename with an equipment row."""
        workbook = self.get_workbook()
        sheet = self.get_equipments_sheet(workbook)
        row_idx = row[0].row
        sheet.cell(row=row_idx, column=EQ_COL_DATASHEET, value=str(filename).strip())
        if not self.save_workbook(workbook):
            return False, "Close stock.xlsx in Excel before saving."
        label = filename.strip() or "(none)"
        return True, f"Datasheet linked: {label}"

    @staticmethod
    def parse_equipment_images(value) -> list[str]:
        text = str(value or "").strip()
        if not text:
            return []
        if EQUIPMENT_IMAGE_SEPARATOR in text:
            return [
                part.strip()
                for part in text.split(EQUIPMENT_IMAGE_SEPARATOR)
                if part.strip()
            ]
        return [text]

    @staticmethod
    def format_equipment_images(filenames: list[str]) -> str:
        names: list[str] = []
        seen: set[str] = set()
        for name in filenames:
            clean = str(name or "").strip()
            if not clean or clean in seen:
                continue
            seen.add(clean)
            names.append(clean)
        return EQUIPMENT_IMAGE_SEPARATOR.join(names)

    def set_equipment_images(self, row, filenames: list[str]) -> tuple[bool, str]:
        workbook = self.get_workbook()
        sheet = self.get_equipments_sheet(workbook)
        row_idx = row[0].row
        sheet.cell(
            row=row_idx,
            column=EQ_COL_IMAGE,
            value=self.format_equipment_images(filenames),
        )
        if not self.save_workbook(workbook):
            return False, "Close stock.xlsx in Excel before saving."
        count = len(self.parse_equipment_images(sheet.cell(row=row_idx, column=EQ_COL_IMAGE).value))
        if count == 0:
            return True, "Equipment images cleared."
        if count == 1:
            return True, "Equipment image linked."
        return True, f"{count} equipment images linked."

    def append_equipment_image(self, row, filename: str) -> tuple[bool, str]:
        name = str(filename or "").strip()
        if not name:
            return False, "Image filename is required."
        current = self.parse_equipment_images(
            self.equipment_row_to_dict(row).get("image", "")
        )
        if name in current:
            return self.set_equipment_images(row, current)
        current.append(name)
        return self.set_equipment_images(row, current)

    def remove_equipment_image(self, row, filename: str) -> tuple[bool, str]:
        name = str(filename or "").strip()
        current = self.parse_equipment_images(
            self.equipment_row_to_dict(row).get("image", "")
        )
        if name not in current:
            return True, "Image removed."
        updated = [item for item in current if item != name]
        return self.set_equipment_images(row, updated)

    def link_equipment_image(self, row, filename: str) -> tuple[bool, str]:
        """Associate one equipment image (appends to the linked list)."""
        return self.append_equipment_image(row, filename)

    def unlink_equipment_image(self, row) -> tuple[bool, str]:
        """Remove all linked equipment images from the row."""
        return self.set_equipment_images(row, [])

    def set_equipment_location(
        self, row, location: str, *, user: str = ""
    ) -> tuple[bool, str]:
        location = str(location).strip()
        workbook = self.get_workbook()
        sheet = self.get_equipments_sheet(workbook)
        row_idx = row[0].row
        data = self.equipment_row_to_dict(row)
        if data.get("loaned"):
            return False, "Return the equipment before changing home location."
        sheet.cell(row=row_idx, column=EQ_COL_LOCATION, value=location)
        if not self.save_workbook(workbook):
            return False, "Close stock.xlsx in Excel before saving."
        return True, "Location updated."

    def loan_equipment_out(
        self,
        row,
        *,
        user: str,
        loaned_to: str,
        place: str,
        home_location: str = "",
        notes: str = "",
    ) -> tuple[bool, str]:
        loaned_to = str(loaned_to).strip()
        place = str(place).strip()
        home_location = str(home_location).strip()
        notes = str(notes).strip()
        if not loaned_to:
            return False, "Provide who received the equipment."
        if not place:
            return False, "Provide where the equipment is (Lab, Factory, etc.)."

        workbook = self.get_workbook()
        sheet = self.get_equipments_sheet(workbook)
        row_idx = row[0].row
        data = self.equipment_row_to_dict(row)
        eq_id = data["id"]
        if not home_location:
            home_location = str(data.get("location", "")).strip()

        since = self._loan_datetime_now()
        sheet.cell(row=row_idx, column=EQ_COL_LOCATION, value=home_location)
        sheet.cell(row=row_idx, column=EQ_COL_LOANED, value="Yes")
        sheet.cell(row=row_idx, column=EQ_COL_LOANED_TO, value=loaned_to)
        sheet.cell(row=row_idx, column=EQ_COL_LOAN_PLACE, value=place)
        sheet.cell(row=row_idx, column=EQ_COL_LOAN_SINCE, value=since)
        self._append_equipment_loan_record(
            workbook,
            user=user,
            equipment_id=eq_id,
            action="OUT",
            loaned_to=loaned_to,
            place=place,
            home_location=home_location,
            notes=notes,
        )
        if not self.save_workbook(workbook):
            return False, "Close stock.xlsx in Excel before saving."
        return True, f"Loaned to {loaned_to} @ {place} ({since})."

    def return_equipment_loan(
        self, row, *, user: str, notes: str = ""
    ) -> tuple[bool, str]:
        workbook = self.get_workbook()
        sheet = self.get_equipments_sheet(workbook)
        row_idx = row[0].row
        data = self.equipment_row_to_dict(row)
        if not data.get("loaned"):
            return True, "Equipment is not on loan."

        eq_id = data["id"]
        home_location = str(data.get("location", "")).strip()
        self._append_equipment_loan_record(
            workbook,
            user=user,
            equipment_id=eq_id,
            action="IN",
            loaned_to=str(data.get("loaned_to", "")).strip(),
            place=str(data.get("loan_place", "")).strip(),
            home_location=home_location,
            notes=notes,
        )
        sheet.cell(row=row_idx, column=EQ_COL_LOANED, value="No")
        sheet.cell(row=row_idx, column=EQ_COL_LOANED_TO, value="")
        sheet.cell(row=row_idx, column=EQ_COL_LOAN_PLACE, value="")
        sheet.cell(row=row_idx, column=EQ_COL_LOAN_SINCE, value="")
        if not self.save_workbook(workbook):
            return False, "Close stock.xlsx in Excel before saving."
        label = home_location or "home"
        return True, f"Equipment returned to {label}."

    def get_equipment_recent_rows(
        self,
        workbook,
        *,
        limit: int = 20,
        equipment_only: bool = False,
        supplier_reference: str = "",
        description: str = "",
    ) -> list:
        """Last N equipment sheet rows (newest last in Excel → reversed)."""
        sheet = self.get_equipments_sheet(workbook)
        ref_filter = self.normalize_ref(supplier_reference) if equipment_only else ""
        desc_filter = self.normalize_ref(description) if equipment_only else ""
        rows: list = []
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            if self.row_is_empty(row):
                continue
            if equipment_only and (ref_filter or desc_filter):
                row_ref = self.normalize_ref(row[1].value)
                row_name = self.normalize_ref(row[3].value)
                row_desc = self.normalize_ref(row[4].value)
                if ref_filter and row_ref != ref_filter:
                    continue
                if desc_filter and desc_filter not in row_name and desc_filter not in row_desc:
                    continue
            rows.append(row)
        if limit > 0:
            rows = rows[-limit:]
        return list(reversed(rows))

    def get_equipment_rows(
        self,
        workbook,
        equipment_only: bool = False,
        supplier_reference: str = "",
        description: str = "",
    ) -> list:
        """Rows for equipments table dialog (last 20, newest first)."""
        sheet = self.get_equipments_sheet(workbook)
        ref_filter = self.normalize_ref(supplier_reference) if equipment_only else ""
        desc_filter = self.normalize_ref(description) if equipment_only else ""
        rows: list[tuple] = []
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            if self.row_is_empty(row):
                continue
            if equipment_only and (ref_filter or desc_filter):
                row_ref = self.normalize_ref(row[1].value)
                row_name = self.normalize_ref(row[3].value)
                row_desc = self.normalize_ref(row[4].value)
                if ref_filter and row_ref != ref_filter:
                    continue
                if desc_filter and desc_filter not in row_name and desc_filter not in row_desc:
                    continue
            data = self.equipment_row_to_dict(row)
            rows.append((
                data["id"],
                data["supplier_reference"],
                data["serial_number"],
                data["name"],
                data["description"],
                data["calibration_date"],
                data["calibration_expiration"],
            ))
        return list(reversed(rows[-20:]))
